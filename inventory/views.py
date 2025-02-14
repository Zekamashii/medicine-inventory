import csv
import datetime
import io
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth import views as auth_views
from django.contrib.auth.views import LogoutView
from django.forms import modelformset_factory
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views.generic import TemplateView, View, CreateView, UpdateView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from pykakasi import kakasi
from tablib import Dataset
from inventory.csv_download import generate_transactions_csv
from inventory.databar_analyzer import analyzer
from inventory.forms import UserRegisterForm, CreateInboundTransactionForm, \
    CreateOutboundTransactionForm, CreateTransferTransactionForm, CreateDrugMasterForm, \
    EditDrugMasterForm, CreateCategoryForm, CreateSiteForm, EditCategoryForm, EditSiteForm, \
    EditUserForm, CreateAdjustTransactionForm, CancelTransactionForm, SafetyStockForm, \
    CustomPasswordChangeForm, DrugObsoleteStatusForm
from inventory.models import InventoryItem, Category, Site, Transaction, Drug, UserProfile, \
    SafetyStock, InventoryInspection, DrugObsoleteBySite
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError
from urllib.parse import quote
from django.db.models import Case, When, Sum, Value, IntegerField, F, Q, Prefetch, Subquery, OuterRef, Min
from django.core.paginator import (
    Paginator,
    EmptyPage,
    PageNotAnInteger,
)
from inventory.resources import DrugMasterResource
from django.views.decorators.csrf import requires_csrf_token
from openpyxl import Workbook
from openpyxl.styles import Font


class Index(TemplateView):
    template_name = 'inventory/index.html'


class SignUpView(View):
    def get(self, request):
        form = UserRegisterForm()
        return render(request, 'inventory/signup.html', {'form': form})

    def post(self, request):
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            user = authenticate(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password1']
            )

            login(request, user)
            return redirect('index')

        return render(request, 'inventory/signup.html', {'form': form})


class CustomLogoutView(LogoutView):
    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        response.delete_cookie('selectedSite')
        return response


class Dashboard(LoginRequiredMixin, View):
    def get(self, request):
        selected_site_color = ''
        selected_site_id = ''
        try:
            selected_site_id = request.COOKIES.get('selectedSite')
            if selected_site_id is None:
                selected_site_name = '全拠点'
                selected_site_color = ''
            else:
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
        except Site.DoesNotExist:
            selected_site_name = '全拠点'

        # Get user profile
        user_profile = UserProfile.objects.get(user=request.user)

        # Sorting
        default_sort_field = user_profile.default_sort_field
        default_sort_order = user_profile.default_sort_order

        sort_field = request.GET.get('sort', default_sort_field)
        sort_order = request.GET.get('order', default_sort_order)

        # Capture the search query
        search_query = request.GET.get('search_query', '')
        selected_category = request.GET.get('category', '')

        # Filtering items based on site selection
        items_filter = {'quantity__gt': 0}
        if selected_site_id is not None:
            items_filter['site'] = selected_site_id
        if selected_category:
            items_filter['name__category_id'] = selected_category

        items = InventoryItem.objects.select_related('name', 'name__category').filter(
            **items_filter)

        if search_query:
            items = items.filter(
                Q(id__iexact=search_query) |
                Q(name__name__icontains=search_query) |
                Q(lot__icontains=search_query) |
                Q(name__product_code__icontains=search_query)
            )
        # Summary calculations
        total_items = items.count()

        # Calculate expiring items
        now = timezone.now()
        date_30_days_from_now = timezone.now() + datetime.timedelta(days=30)

        expiring_items_query = items.filter(expire_date__gt=now,
                                            expire_date__lte=date_30_days_from_now)
        expired_items_query = items.filter(expire_date__lt=now)
        expire_inventory_ids = list(expiring_items_query.values_list('id', flat=True))
        expired_inventory_ids = list(expired_items_query.values_list('id', flat=True))
        expire_count = expiring_items_query.count()
        expired_count = expired_items_query.count()

        # Calculate low inventory items.
        aggregated_inventory = (
            InventoryItem.objects.filter(**items_filter)
            .values('name', 'site')
            .annotate(
                total_quantity=Sum('quantity'),
            )
        )

        # Join with SafetyStock to get the correct min_stock per drug and site
        aggregated_inventory = aggregated_inventory.annotate(
            min_stock=Subquery(
                SafetyStock.objects.filter(
                    drug_id=OuterRef('name'),
                    site_id=OuterRef('site')
                ).values('min_stock')[:1]
            )
        )

        low_inventory_records = aggregated_inventory.filter(
            total_quantity__lt=F('min_stock'),
            name__obsoleted=False  # Add this to filter out obsolete drugs at this stage
        )

        all_low_inventory_ids = []
        for record in low_inventory_records:
            drug_id = record['name']
            site_id = record['site']
            low_inventory_items = InventoryItem.objects.filter(
                name_id=drug_id,
                site_id=site_id,
                name__obsoleted=False
            ).values_list('id', flat=True)

            all_low_inventory_ids.extend(low_inventory_items)

        low_inventory_ids = all_low_inventory_ids
        low_inventory_count = len(low_inventory_ids)

        # Apply the filter based on the clicked card
        filter_type = request.GET.get('filter', '')

        if filter_type == 'low_inventory':
            items = items.filter(id__in=low_inventory_ids)
        elif filter_type == 'expiring':
            items = items.filter(id__in=expire_inventory_ids)
        elif filter_type == 'expired':
            items = items.filter(id__in=expired_inventory_ids)

        # Apply custom sort order if sorting by name
        if sort_field == 'name':
            items = items.annotate(
                custom_sort_order=Case(
                    When(name__kana__regex=r'^[ぁ-ん]', then=Value(1)),
                    When(name__kana__regex=r'^[a-zA-Z]', then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )
            ).order_by(
                f"{'-' if sort_order == 'desc' else ''}custom_sort_order",
                f"{'-' if sort_order == 'desc' else ''}name__kana"
            )
        elif sort_field == 'status':
            items = items.annotate(
                status_order=Case(
                    When(id__in=expired_inventory_ids, then=Value(1)),
                    When(id__in=expire_inventory_ids, then=Value(2)),
                    When(id__in=low_inventory_ids, then=Value(3)),
                    default=Value(4),
                    output_field=IntegerField(),
                )
            ).order_by(f'{"-" if sort_order == "desc" else ""}status_order')
        else:
            sort_field = f'-{sort_field}' if sort_order == 'desc' else sort_field
            items = items.order_by(sort_field)

        default_page = 1
        page = request.GET.get('page', default_page)
        items_per_page = 50
        paginator = Paginator(items, items_per_page)

        today = datetime.date.today()
        for item in items:
            if item.expire_date:
                remaining_days = (item.expire_date - today).days
                item.remaining_shelf_life = remaining_days if remaining_days >= 0 else '期限切れ'
            else:
                item.remaining_shelf_life = 'N/A'
        try:
            items_page = paginator.page(page)
        except PageNotAnInteger:
            items_page = paginator.page(default_page)
        except EmptyPage:
            items_page = paginator.page(paginator.num_pages)

        low_inventory_percentage = round((low_inventory_count / total_items * 100)) if total_items else 0
        expire_percentage = round((expire_count / total_items * 100)) if total_items else 0
        expired_percentage = round((expired_count / total_items * 100)) if total_items else 0

        page_in_path = 'page' in request.GET

        context = {
            'items': items_page,
            'sites': Site.objects.filter(obsoleted=False),
            'selected_site_id': selected_site_id,
            'selected_site_name': selected_site_name,
            'selected_site_color': selected_site_color,
            'total_items': total_items,
            'low_inventory_ids': low_inventory_ids,
            'low_inventory_count': low_inventory_count,
            'low_inventory_percentage': low_inventory_percentage,
            'expire_count': expire_count,
            'expire_inventory_ids': expire_inventory_ids,
            'expire_percentage': expire_percentage,
            'expired_count': expired_count,
            'expired_inventory_ids': expired_inventory_ids,
            'expired_percentage': expired_percentage,
            'search_query': search_query,
            'categories': Category.objects.filter(obsoleted=False),
            'selected_category': request.GET.get('category', ''),
            'default_sort_field': default_sort_field,
            'default_sort_order': default_sort_order,
            'page_in_path': page_in_path,
            'current_sort': sort_field,
            'current_order': sort_order
        }

        return render(request, 'inventory/dashboard.html', context)


class TransactionsView(LoginRequiredMixin, View):
    def get(self, request):
        selected_site_color = ''
        selected_site_id = ''
        try:
            selected_site_id = request.COOKIES.get('selectedSite')
            if selected_site_id is None:
                selected_site_name = '全拠点'
                selected_site_color = ''
            else:
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
        except Site.DoesNotExist:
            selected_site_name = '全拠点'

        default_sort_field = 'date_created'
        default_sort_order = 'desc'

        sort_field = request.GET.get('sort', default_sort_field)
        sort_order = request.GET.get('order', default_sort_order)

        if sort_order == 'desc':
            sort_field = f'-{sort_field}'

        if selected_site_id is not None:
            items = Transaction.objects.filter(
                Q(source_site=selected_site_id) | Q(dest_site=selected_site_id)
            )
        else:
            items = Transaction.objects.all()

        # Capture the search query
        search_query = request.GET.get('search_query', '')
        if search_query:
            items = items.filter(
                Q(id__iexact=search_query) |
                Q(name__name__icontains=search_query) |
                Q(lot__icontains=search_query) |
                Q(name__product_code__icontains=search_query)
            )

        # Apply custom sort order if sorting by name
        if 'name' in sort_field:
            items = items.annotate(
                custom_sort_order=Case(
                    When(name__kana__regex=r'^[ぁ-ん]', then=Value(1)),
                    When(name__kana__regex=r'^[a-zA-Z]', then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )
            ).order_by(
                f"{'-' if sort_order == 'desc' else ''}custom_sort_order",
                f"{'-' if sort_order == 'desc' else ''}name__kana"
            )
        else:
            items = items.order_by(sort_field)

        default_page = 1
        page = request.GET.get('page', default_page)
        items_per_page = 50
        paginator = Paginator(items, items_per_page)

        try:
            items_page = paginator.page(page)
        except PageNotAnInteger:
            items_page = paginator.page(default_page)
        except EmptyPage:
            items_page = paginator.page(paginator.num_pages)

        context = {
            'items': items_page,
            'sites': Site.objects.filter(obsoleted=False),
            'selected_site_name': selected_site_name,
            'selected_site_color': selected_site_color,
            'search_query': search_query,
            'default_sort_field': default_sort_field,
            'default_sort_order': default_sort_order,
            'current_sort': sort_field,
            'current_order': sort_order
        }

        return render(request, 'inventory/transaction.html', context)


class CategoryView(LoginRequiredMixin, View):
    def get(self, request):
        sort_field = request.GET.get('sort', 'id')
        sort_order = request.GET.get('order', 'asc')
        sort_field = f'-{sort_field}' if sort_order == 'desc' else sort_field

        items = Category.objects.all().order_by(sort_field)

        context = {
            'items': items
        }

        return render(request, 'inventory/category.html', context)


class CreateCategoryView(LoginRequiredMixin, CreateView):
    model = Category
    form_class = CreateCategoryForm
    template_name = 'inventory/category_form.html'
    success_url = reverse_lazy('categories')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class EditCategoryView(LoginRequiredMixin, UpdateView):
    model = Category
    form_class = EditCategoryForm
    template_name = 'inventory/category_form.html'
    success_url = reverse_lazy('categories')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.get_object()
        return context


class SiteView(LoginRequiredMixin, View):
    def get(self, request):
        sort_field = request.GET.get('sort', 'id')
        sort_order = request.GET.get('order', 'asc')
        sort_field = f'-{sort_field}' if sort_order == 'desc' else sort_field

        items = Site.objects.all().order_by(sort_field)

        context = {
            'items': items
        }

        return render(request, 'inventory/site.html', context)


class CreateSiteView(LoginRequiredMixin, CreateView):
    model = Site
    form_class = CreateSiteForm
    template_name = 'inventory/site_form.html'
    success_url = reverse_lazy('sites')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class EditSiteView(LoginRequiredMixin, UpdateView):
    model = Site
    form_class = EditSiteForm
    template_name = 'inventory/site_form.html'
    success_url = reverse_lazy('sites')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.get_object()
        return context


class UserView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        # This method must return True for the view to be accessible.
        return self.request.user.is_superuser

    def handle_no_permission(self):
        response = render(self.request, 'inventory/errors/403.html')
        response.status_code = 403
        return response

    def get(self, request):
        items = User.objects.all().select_related('userprofile')

        context = {
            'items': items
        }

        return render(request, 'inventory/user.html', context)


class EditUserView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = UserProfile
    form_class = EditUserForm
    template_name = 'inventory/user_form.html'
    success_url = reverse_lazy('dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_profile = self.get_object()
        context['items'] = user_profile
        context['first_name'] = user_profile.user.first_name
        context['last_name'] = user_profile.user.last_name
        context['username'] = user_profile.user.username
        context['userid'] = user_profile.user.id
        context['email'] = user_profile.user.email
        context['date_joined'] = user_profile.user.date_joined
        context['last_login'] = user_profile.user.last_login
        return context

    def test_func(self):
        user_profile = self.get_object()
        return user_profile.user == self.request.user


class CustomPasswordChangeView(auth_views.PasswordChangeView):
    form_class = CustomPasswordChangeForm
    template_name = 'inventory/password_change_form.html'

    def get_success_url(self):
        return reverse('edit-user', kwargs={'pk': self.request.user.pk})


class DrugMasterView(LoginRequiredMixin, View):
    def get(self, request):
        default_sort_field = 'id'
        default_sort_order = 'asc'
        sort_field = request.GET.get('sort', default_sort_field)
        sort_order = request.GET.get('order', default_sort_order)
        is_name_sort = sort_field == 'name'

        if is_name_sort:
            drugs = Drug.objects.annotate(
                custom_sort_order=Case(
                    When(kana__regex=r'^[ぁ-ん]', then=Value(1)),
                    When(kana__regex=r'^[a-zA-Z]', then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )
            )
            if sort_order == 'desc':
                drugs = drugs.order_by('-custom_sort_order', '-kana')
            else:
                drugs = drugs.order_by('custom_sort_order', 'kana')
        else:
            if sort_order == 'desc':
                sort_field = f'-{sort_field}'
            drugs = Drug.objects.all().order_by(sort_field)

        search_query = request.GET.get('search_query', '')
        if search_query:
            drugs = drugs.filter(
                Q(id__iexact=search_query) |
                Q(name__icontains=search_query) |
                Q(kana__icontains=search_query) |
                Q(product_code__icontains=search_query)
            )

        default_page = 1
        page = request.GET.get('page', default_page)
        items_per_page = 100
        paginator = Paginator(drugs, items_per_page)

        try:
            items_page = paginator.page(page)
        except PageNotAnInteger:
            items_page = paginator.page(default_page)
        except EmptyPage:
            items_page = paginator.page(paginator.num_pages)

        context = {
            'items': items_page,
            'default_sort_field': default_sort_field,
            'default_sort_order': default_sort_order
        }

        return render(request, 'inventory/drug_master.html', context)


class CreateDrugMasterView(LoginRequiredMixin, CreateView):
    model = Drug
    form_class = CreateDrugMasterForm
    template_name = 'inventory/drug_master_form.html'
    success_url = reverse_lazy('drugs')

    def post(self, request, *args, **kwargs):
        if 'action' in request.POST and request.POST['action'] == 'add_more':
            self.object = None
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form, add_more=True)
            else:
                return self.form_invalid(form)
        else:
            return super().post(request, *args, **kwargs)

    def form_valid(self, form, add_more=False):
        form.instance.user = self.request.user
        self.object = form.save()
        if add_more:
            return HttpResponseRedirect(reverse_lazy('create-drug-master'))
        else:
            return HttpResponseRedirect(self.success_url)


class EditDrugMasterView(LoginRequiredMixin, UpdateView):
    model = Drug
    form_class = EditDrugMasterForm
    template_name = 'inventory/drug_master_form.html'
    success_url = reverse_lazy('drugs')

    def form_valid(self, form):
        form.instance.user = self.request.user
        context = self.get_context_data()
        safety_stock_formset = context['safety_stock_formset']
        obsolete_status_formset = context['obsolete_status_formset']

        if form.has_changed() and 'name' in form.changed_data:
            kks = kakasi()
            kks.setMode("J", "H")  # Kanji to Hiragana
            kks.setMode('K', 'H')  # Katakana to Hiragana
            conv = kks.getConverter()
            form.instance.kana = conv.do(form.instance.name)

        if form.is_valid() and safety_stock_formset.is_valid() and obsolete_status_formset.is_valid():
            response = super().form_valid(form)
            safety_stock_formset.save()
            obsolete_status_formset.save()
            return response
        else:
            return self.render_to_response(self.get_context_data(
                form=form,
                safety_stock_formset=safety_stock_formset,
                obsolete_status_formset=obsolete_status_formset
            ))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.get_object()

        safety_stock_form_set = modelformset_factory(SafetyStock, form=SafetyStockForm, extra=0)
        obsolete_status_form_set = modelformset_factory(DrugObsoleteBySite,
                                                        form=DrugObsoleteStatusForm, extra=0)

        if 'safety_stock_formset' in kwargs:
            context['safety_stock_formset'] = kwargs['safety_stock_formset']
        else:
            context['safety_stock_formset'] = safety_stock_form_set(
                self.request.POST or None,
                queryset=SafetyStock.objects.filter(drug=context['item'], site__obsoleted=False)
            )

        if 'obsolete_status_formset' in kwargs:
            context['obsolete_status_formset'] = kwargs['obsolete_status_formset']
        else:
            context['obsolete_status_formset'] = obsolete_status_form_set(
                self.request.POST or None,
                queryset=DrugObsoleteBySite.objects.filter(drug=context['item'], site__obsoleted=False),
                prefix='obsolete_status'
            )
        return context


class CreateInboundTransactionView(LoginRequiredMixin, CreateView):
    model = Transaction
    form_class = CreateInboundTransactionForm
    template_name = 'inventory/transaction_form.html'
    success_url = reverse_lazy('dashboard')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['drugs'] = Drug.objects.filter(obsoleted=False).annotate(
            custom_sort_order=Case(
                When(kana__regex=r'^[ぁ-ん]', then=Value(1)),
                When(kana__regex=r'^[a-zA-Z]', then=Value(2)),
                default=Value(3),
                output_field=IntegerField(),
            )
        ).order_by(
            'custom_sort_order',
            'kana'
        )
        return kwargs

    def post(self, request, *args, **kwargs):
        if 'action' in request.POST and request.POST['action'] == 'add_more':
            self.object = None
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form, add_more=True)
            else:
                return self.form_invalid(form)
        else:
            return super().post(request, *args, **kwargs)

    def form_valid(self, form, add_more=False):
        try:
            self.check_obsolete(form.instance)
            self.check_conflict(form.instance)
        except ValidationError as e:
            form.add_error(None, e.message)
            return self.form_invalid(form)

        with transaction.atomic():
            form.instance.user = self.request.user
            transaction_instance = form.save()
            self.create_inventory(transaction_instance)

        if add_more:
            return HttpResponseRedirect(reverse_lazy('create_inbound_transaction'))
        else:
            return HttpResponseRedirect(self.success_url)

    def form_invalid(self, form):
        errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
        return JsonResponse({'errors': errors}, status=400)

    def check_obsolete(self, instance):
        new_items = Drug.objects.filter(name=instance.name)
        for item in new_items:
            if item.obsoleted:
                raise ValidationError(
                    f"{item.name}のマスターは廃止されたとマークされています。「クリア」ボタンを押して、再度入力してください。")

    def check_conflict(self, instance):
        conflicting_items = InventoryItem.objects.filter(
            name=instance.name,
            lot=instance.lot,
        ).exclude(expire_date=instance.expire_date)

        if conflicting_items.exists():
            raise ValidationError(
                f"同じ製造番号を持つ薬は同じ有効期限でなければなりません。使用期限を再度入力してください。")

    def create_inventory(self, instance):
        inventory_items = InventoryItem.objects.filter(
            name=instance.name,
            unit=instance.unit,
            expire_date=instance.expire_date,
            lot=instance.lot,
            site=instance.dest_site
        )

        if inventory_items.exists():
            for item in inventory_items:
                item.quantity += instance.quantity
                item.user = instance.user
                item.save()
        else:
            InventoryItem.objects.create(
                name=instance.name,
                quantity=instance.quantity,
                unit=instance.unit,
                lot=instance.lot,
                expire_date=instance.expire_date,
                site=instance.dest_site,
                user=instance.user
            )

    def get_selected_site_info(self):
        selected_site_id = self.request.COOKIES.get('selectedSite')
        selected_site_name = '全拠点'
        selected_site_color = ''

        if selected_site_id:
            try:
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
            except Site.DoesNotExist:
                pass

        return selected_site_name, selected_site_color

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_site_name, selected_site_color = self.get_selected_site_info()
        context['type'] = Transaction.Type.INBOUND.name
        context['type_label'] = Transaction.Type.INBOUND.label
        context['switch_name'] = Transaction.Type.OUTBOUND.label
        context['switch_url'] = 'create_outbound_transaction'
        context['form_action_url'] = reverse_lazy('create_inbound_transaction')
        context['selected_site_name'] = selected_site_name
        context['selected_site_color'] = selected_site_color
        return context

    def get_initial(self):
        initial = super().get_initial()
        initial['type'] = Transaction.Type.INBOUND
        selected_site_id = self.request.COOKIES.get('selectedSite')
        scanner_mode_enabled = self.request.user.userprofile.scanner_mode_enabled
        initial['scanner_mode_enabled'] = scanner_mode_enabled
        if selected_site_id and selected_site_id.isdigit():
            initial['dest_site'] = selected_site_id
        return initial


class CreateOutboundTransactionView(LoginRequiredMixin, CreateView):
    model = Transaction
    form_class = CreateOutboundTransactionForm
    template_name = 'inventory/transaction_form.html'
    success_url = reverse_lazy('dashboard')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['drugs'] = Drug.objects.filter(obsoleted=False).annotate(
            custom_sort_order=Case(
                When(kana__regex=r'^[ぁ-ん]', then=Value(1)),
                When(kana__regex=r'^[a-zA-Z]', then=Value(2)),
                default=Value(3),
                output_field=IntegerField(),
            )
        ).order_by(
            'custom_sort_order',
            'kana'
        )
        return kwargs

    def post(self, request, *args, **kwargs):
        if 'action' in request.POST and request.POST['action'] == 'add_more':
            self.object = None
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form, add_more=True)
            else:
                return self.form_invalid(form)
        else:
            return super().post(request, *args, **kwargs)

    def form_valid(self, form, add_more=False):
        try:
            self.check_inventory_levels(form.instance)
            self.check_lot(form.instance)
        except ValidationError as e:
            form.add_error(None, e.message)
            return self.form_invalid(form)

        with transaction.atomic():
            form.instance.user = self.request.user
            transaction_instance = form.save()
            self.adjust_inventory(transaction_instance)

        if add_more:
            return HttpResponseRedirect(reverse_lazy('create_outbound_transaction'))
        else:
            return HttpResponseRedirect(self.success_url)

    def form_invalid(self, form):
        errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
        return JsonResponse({'errors': errors}, status=400)

    def check_inventory_levels(self, instance):
        inventory_items = InventoryItem.objects.filter(
            name=instance.name,
            unit=instance.unit,
            lot=instance.lot,
            site=instance.source_site
        )

        for item in inventory_items:
            if item.quantity < instance.quantity:
                raise ValidationError(
                    f"{item.name}の在庫が不足しています。必要数: {instance.quantity}、在庫数: {item.quantity}。取引数量を再度入力してください。")

    def check_lot(self, instance):
        matching_lots = InventoryItem.objects.filter(name=instance.name, unit=instance.unit,
                                                     site=instance.source_site)
        lot_match = any(item.lot == instance.lot for item in matching_lots)
        if not lot_match:
            raise ValidationError(
                f"{instance.name}の製造番号が不正です。「クリア」ボタンを押して、再度入力してください。")

    def adjust_inventory(self, instance):
        inventory_items = InventoryItem.objects.filter(
            name=instance.name,
            unit=instance.unit,
            lot=instance.lot,
            site=instance.source_site
        )

        for item in inventory_items:
            item.quantity -= instance.quantity
            item.user = instance.user
            item.date_created = instance.date_created
            item.save()

    def get_selected_site_info(self):
        selected_site_id = self.request.COOKIES.get('selectedSite')
        selected_site_name = '全拠点'
        selected_site_color = ''

        if selected_site_id:
            try:
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
            except Site.DoesNotExist:
                pass

        return selected_site_name, selected_site_color

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_site_name, selected_site_color = self.get_selected_site_info()
        context['type'] = Transaction.Type.OUTBOUND.name
        context['type_label'] = Transaction.Type.OUTBOUND.label
        context['switch_name'] = Transaction.Type.INBOUND.label
        context['switch_url'] = 'create_inbound_transaction'
        context['form_action_url'] = reverse_lazy('create_outbound_transaction')
        context['selected_site_name'] = selected_site_name
        context['selected_site_color'] = selected_site_color
        return context

    def get_initial(self):
        initial = super().get_initial()
        initial['type'] = Transaction.Type.OUTBOUND
        selected_site_id = self.request.COOKIES.get('selectedSite')
        scanner_mode_enabled = self.request.user.userprofile.scanner_mode_enabled
        initial['scanner_mode_enabled'] = scanner_mode_enabled
        if selected_site_id and selected_site_id.isdigit():
            initial['source_site'] = selected_site_id
        return initial


class CreateTransferTransactionView(LoginRequiredMixin, CreateView):
    model = Transaction
    form_class = CreateTransferTransactionForm
    template_name = 'inventory/transaction_form.html'
    success_url = reverse_lazy('dashboard')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['drugs'] = Drug.objects.filter(obsoleted=False).annotate(
            custom_sort_order=Case(
                When(kana__regex=r'^[ぁ-ん]', then=Value(1)),
                When(kana__regex=r'^[a-zA-Z]', then=Value(2)),
                default=Value(3),
                output_field=IntegerField(),
            )
        ).order_by(
            'custom_sort_order',
            'kana'
        )
        return kwargs

    def post(self, request, *args, **kwargs):
        if 'action' in request.POST and request.POST['action'] == 'add_more':
            self.object = None
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form, add_more=True)
            else:
                return self.form_invalid(form)
        else:
            return super().post(request, *args, **kwargs)

    def form_valid(self, form, add_more=False):
        try:
            self.check_inventory_levels(form.instance)
            self.check_lot(form.instance)
            self.site_check(form.instance)
        except ValidationError as e:
            form.add_error(None, e.message)
            return self.form_invalid(form)

        with transaction.atomic():
            form.instance.user = self.request.user
            form.instance.status = Transaction.Status.PENDING
            form.instance.source_inventory_adjusted = True
            transaction_instance = form.save()
            self.adjust_inventory(transaction_instance)

        if add_more:
            return HttpResponseRedirect(reverse_lazy('create_transfer_transaction'))
        else:
            return HttpResponseRedirect(self.success_url)

    def form_invalid(self, form):
        errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
        return JsonResponse({'errors': errors}, status=400)

    def site_check(self, instance):
        if instance.source_site == instance.dest_site:
            raise ValidationError("出庫元と入庫先が同一ではないようにしてください。")

    def check_lot(self, instance):
        matching_lots = InventoryItem.objects.filter(name=instance.name, unit=instance.unit,
                                                     site=instance.source_site)
        lot_match = any(item.lot == instance.lot for item in matching_lots)
        if not lot_match:
            raise ValidationError(
                f"{instance.name}の製造番号が不正です。「クリア」ボタンを押して、再度入力してください。")

    def check_inventory_levels(self, instance):
        inventory_items = InventoryItem.objects.filter(name=instance.name, unit=instance.unit,
                                                       lot=instance.lot, site=instance.source_site)
        for item in inventory_items:
            if item.quantity < instance.quantity:
                raise ValidationError(
                    f"{item.name}の在庫が不足しています。必要数: {instance.quantity}、在庫数: {item.quantity}。取引数量を再度入力してください。")

    def adjust_inventory(self, instance):
        source_inventory_items = InventoryItem.objects.filter(name=instance.name,
                                                              unit=instance.unit, lot=instance.lot,
                                                              site=instance.source_site)
        for item in source_inventory_items:
            item.quantity -= instance.quantity
            item.user = instance.user
            item.date_created = instance.date_created
            item.save()

    def get_selected_site_info(self):
        selected_site_id = self.request.COOKIES.get('selectedSite')
        selected_site_name = '全拠点'
        selected_site_color = ''

        if selected_site_id:
            try:
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
            except Site.DoesNotExist:
                pass

        return selected_site_name, selected_site_color

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_site_name, selected_site_color = self.get_selected_site_info()
        context['type'] = Transaction.Type.TRANSFER.name
        context['type_label'] = Transaction.Type.TRANSFER.label
        context['switch_name'] = Transaction.Type.TRANSFER.label
        context['switch_url'] = "create_transfer_transaction"
        context['form_action_url'] = reverse_lazy('create_transfer_transaction')
        context['selected_site_name'] = selected_site_name
        context['selected_site_color'] = selected_site_color
        return context

    def get_initial(self):
        initial = super().get_initial()
        initial['type'] = Transaction.Type.TRANSFER
        selected_site_id = self.request.COOKIES.get('selectedSite')
        scanner_mode_enabled = self.request.user.userprofile.scanner_mode_enabled
        initial['scanner_mode_enabled'] = scanner_mode_enabled
        if selected_site_id and selected_site_id.isdigit():
            initial['source_site'] = selected_site_id
        return initial


class ConfirmTransferView(LoginRequiredMixin, UpdateView):
    model = Transaction
    fields = []
    template_name = 'inventory/confirm_transfer.html'
    success_url = reverse_lazy('dashboard')

    def get_object(self, queryset=None):
        obj = get_object_or_404(Transaction, pk=self.kwargs['pk'], type=Transaction.Type.TRANSFER)
        return obj

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            with transaction.atomic():
                self.object.status = Transaction.Status.RECEIVED
                self.object.save()
                self.adjust_inventory(self.object)
                return HttpResponseRedirect(self.get_success_url())
        except ValidationError as e:
            messages.error(request, str(e))
            return self.render_to_response(self.get_context_data())

    def adjust_inventory(self, instance):
        try:
            dest_inventory_item = InventoryItem.objects.get(
                name=instance.name,
                unit=instance.unit,
                lot=instance.lot,
                site=instance.dest_site
            )
        except InventoryItem.DoesNotExist:
            dest_inventory_item = InventoryItem(
                name=instance.name,
                unit=instance.unit,
                lot=instance.lot,
                site=instance.dest_site,
                user=instance.user,
                quantity=0
            )

        dest_inventory_item.expire_date = instance.expire_date
        dest_inventory_item.quantity += instance.quantity
        dest_inventory_item.user = self.request.user
        dest_inventory_item.save()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.get_object()
        return context


class PendingConfirmationsView(LoginRequiredMixin, ListView):
    model = Transaction
    template_name = 'inventory/pending_confirmations.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = Transaction.objects.filter(
            type=Transaction.Type.TRANSFER,
            status=Transaction.Status.PENDING
        ).order_by('-date_created')
        return queryset


class CreateAdjustTransactionView(LoginRequiredMixin, CreateView):
    model = Transaction
    form_class = CreateAdjustTransactionForm
    template_name = 'inventory/adjust_form.html'
    success_url = reverse_lazy('dashboard')

    def get_initial(self):
        initial = super().get_initial()
        initial['type'] = Transaction.Type.ADJUST
        item_id = self.request.GET.get('item_id')
        initial['scanner_mode_enabled'] = False
        if item_id:
            inventory_item = InventoryItem.objects.get(id=item_id)
            self.inventory_item_id = inventory_item.id
            initial['name'] = inventory_item.name
            initial['quantity'] = inventory_item.quantity
            initial['unit'] = inventory_item.unit
            initial['lot'] = inventory_item.lot
            initial['source_site'] = inventory_item.site
            initial['expire_date'] = inventory_item.expire_date
        return initial

    def form_valid(self, form):
        try:
            self.item_check(form.instance)
        except ValidationError as e:
            form.add_error(None, e.message)
            return self.form_invalid(form)

        with transaction.atomic():
            form.instance.user = self.request.user
            transaction_instance = form.save()
            self.adjust_inventory(transaction_instance)
            return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['type'] = Transaction.Type.ADJUST.name
        context['type_label'] = Transaction.Type.ADJUST.label
        context['switch_name'] = Transaction.Type.ADJUST.label
        context['switch_url'] = "create_adjust_transaction"
        return context

    def item_check(self, instance):
        inventory_items = InventoryItem.objects.filter(
            name=instance.name,
            unit=instance.unit,
            lot=instance.lot,
            site=instance.source_site
        )
        for item in inventory_items:
            if not item:
                raise ValidationError(f"Mismatch")

    def adjust_inventory(self, instance):
        try:
            inventory_item_id = getattr(self, 'inventory_item_id', None)
            if inventory_item_id:
                inventory_item = InventoryItem.objects.get(id=inventory_item_id)
                inventory_item.quantity = instance.quantity
                inventory_item.date_created = instance.date_created
                inventory_item.user = self.request.user
                inventory_item.save()

        except Exception as e:
            print(f"Unexpected error: {e}")


class CancelTransactionView(LoginRequiredMixin, UpdateView):
    model = Transaction
    form_class = CancelTransactionForm
    template_name = 'inventory/cancel_form.html'
    success_url = reverse_lazy('transactions')

    def dispatch(self, request, *args, **kwargs):
        self.transaction_item = get_object_or_404(Transaction, pk=kwargs.get('pk'))
        if self.transaction_item.type in [Transaction.Type.ADJUST, Transaction.Type.CANCELLED]:
            messages.error(request,
                           "調整済みまたはキャンセル済みの取引をキャンセルすることはできません。")
            return redirect('transactions')
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        item_id = self.request.GET.get('item_id')
        if item_id:
            transaction_item = Transaction.objects.get(id=item_id)
            self.transaction_item_id = transaction_item.id
            for field in ['name', 'quantity', 'unit', 'lot', 'source_site', 'dest_site',
                          'expire_date', 'date_created', 'user']:
                initial[field] = getattr(transaction_item, field)
        return initial

    def form_valid(self, form):
        with transaction.atomic():
            form.instance.user = self.request.user
            transaction_instance = form.save()
            self.adjust_inventory(transaction_instance)
            transaction_instance.type = Transaction.Type.CANCELLED
            transaction_instance.date_created = timezone.now()
            transaction_instance.save()
            return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.get_object()
        return context

    def adjust_inventory(self, instance):
        if instance.type == Transaction.Type.INBOUND:
            inventory_item, created = InventoryItem.objects.get_or_create(
                name=instance.name,
                unit=instance.unit,
                lot=instance.lot,
                site=instance.dest_site
            )
            new_quantity = max(inventory_item.quantity - instance.quantity, 0)
            inventory_item.quantity = new_quantity
            inventory_item.save()

        elif instance.type == Transaction.Type.OUTBOUND:
            inventory_item, created = InventoryItem.objects.get_or_create(
                name=instance.name,
                unit=instance.unit,
                lot=instance.lot,
                site=instance.source_site,
                expire_date=instance.expire_date
            )
            inventory_item.quantity = F('quantity') + instance.quantity
            inventory_item.save()

        elif instance.type == Transaction.Type.TRANSFER:
            source_inventory_item, _ = InventoryItem.objects.get_or_create(
                name=instance.name,
                unit=instance.unit,
                lot=instance.lot,
                site=instance.source_site,
                expire_date=instance.expire_date
            )
            source_inventory_item.quantity = F('quantity') + instance.quantity
            source_inventory_item.save()

            dest_inventory_item = InventoryItem.objects.get(
                name=instance.name,
                unit=instance.unit,
                lot=instance.lot,
                site=instance.dest_site
            )
            # Ensure the quantity doesn't drop below zero
            new_quantity = max(dest_inventory_item.quantity - instance.quantity, 0)
            dest_inventory_item.quantity = new_quantity
            dest_inventory_item.save()


class InventoryCalendarView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = timezone.now().date()
        period_days = request.GET.get('period', 28)
        try:
            period_days = int(period_days)
        except ValueError:
            period_days = 28

        historical_period = request.GET.get('historical_period', 28)
        try:
            historical_period = int(historical_period)
        except ValueError:
            historical_period = 28

        some_days_later = today + datetime.timedelta(days=period_days)
        selected_site_color = ''
        selected_site_id = ''

        try:
            selected_site_id = request.COOKIES.get('selectedSite')
            if selected_site_id is None:
                selected_site = ''
                selected_site_name = '全拠点'
                selected_site_color = ''
            else:
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
        except Site.DoesNotExist:
            selected_site_name = '全拠点'

        items_filter = {'quantity__gt': -1}
        if selected_site_id is not None:
            items_filter['site_id'] = selected_site_id

        safety_stock_prefetch = Prefetch(
            'name__safetystock_set',
            queryset=SafetyStock.objects.filter(
                site_id=selected_site_id) if selected_site_id else SafetyStock.objects.all(),
            to_attr='safety_stock'
        )

        if selected_site_id:
            selected_site = Site.objects.get(id=selected_site_id)
            obsolete_drug_ids = DrugObsoleteBySite.objects.filter(
                site=selected_site,
                is_obsolete=True
            ).values_list('drug_id', flat=True)
        else:
            obsolete_drug_ids = []

        # Adjusted query to include SafetyStock data
        inventory_items = InventoryItem.objects.filter(
            **items_filter,
            name__obsoleted=False
        ).exclude(
            name_id__in=obsolete_drug_ids
        ).select_related(
            'name',
            'name__category'
        ).prefetch_related(
            safety_stock_prefetch
        ).order_by(
            'name__category__id',
            'name__id'
        )

        periods = [
            (today + datetime.timedelta(days=i), today + datetime.timedelta(days=i + 6))
            for i in range(0, (some_days_later - today).days + 1, 7)
        ]

        # Get historical usage data
        historical_start_date = today - datetime.timedelta(days=historical_period)
        historical_end_date = today

        transactions = Transaction.objects.filter(
            type='O',
            date_created__gte=historical_start_date,
            date_created__lte=historical_end_date,
        )

        if selected_site_id:
            transactions = transactions.filter(source_site_id=selected_site_id)

        usage_data = transactions.values('name_id').annotate(total_quantity=Sum('quantity'))

        usage_per_week = {}
        for data in usage_data:
            name_id = data['name_id']
            total_quantity = data['total_quantity']
            average_weekly_usage = total_quantity * 7 / historical_period
            usage_per_week[name_id] = average_weekly_usage

        # Build initial quantities per drug
        initial_quantities = {}
        inventory_by_period = {}

        for item in inventory_items:
            name_id = item.name.id
            category = item.name.category.name
            name = item.name.name

            if name_id not in inventory_by_period:
                if selected_site_name == '全拠点':
                    safety_stock = ''
                else:
                    safety_stock = item.name.safety_stock[0].min_stock if item.name.safety_stock else 'N/A'

                inventory_by_period[name_id] = {
                    'category': category,
                    'name': name,
                    'safety_stock': safety_stock,
                    'periods': [
                        {"Period": f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}",
                         "Quantity": 0} for start, end in periods
                    ]
                }

            # Sum up the total quantity per drug
            if name_id not in initial_quantities:
                initial_quantities[name_id] = 0

            if item.expire_date >= today:
                initial_quantities[name_id] += item.quantity

        # Adjust quantities for each period
        for name_id, data in inventory_by_period.items():
            total_quantity = initial_quantities.get(name_id, 0)
            data['periods'][0]['Quantity'] = total_quantity

            average_weekly_usage = usage_per_week.get(name_id, 0)

            for period_index in range(1, len(periods)):
                previous_quantity = data['periods'][period_index - 1]["Quantity"]
                predicted_quantity = previous_quantity - average_weekly_usage
                data['periods'][period_index]["Quantity"] = int(round(predicted_quantity, 0))

        # Calculate difference
        for name_id, data in inventory_by_period.items():
            if selected_site_name == '全拠点':
                data['difference'] = ''
            else:
                safety_stock = data['safety_stock']
                first_period_quantity = data['periods'][0]["Quantity"]
                if isinstance(safety_stock, (int, float)) and safety_stock != 'N/A':
                    data['difference'] = first_period_quantity - safety_stock
                else:
                    data['difference'] = ''

        # Reorganize data into category_items
        category_items = {}
        for name_id, data in inventory_by_period.items():
            category = data['category']
            name = data['name']
            if category not in category_items:
                category_items[category] = {}
            category_items[category][name] = data

        context = {
            'sites': Site.objects.filter(obsoleted=False),
            'start_date': today,
            'end_date': some_days_later,
            'inventory_by_period': category_items,
            'periods': periods,
            'period_days': period_days,
            'historical_period': historical_period,
            'selected_site_name': selected_site_name,
            'selected_site_color': selected_site_color,
            'username': request.user.username,
            'current_time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        return render(request, 'inventory/calendar.html', context)



class InventoryValidationView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            selected_site_id = request.COOKIES.get('selectedSite', None)
            if selected_site_id:
                sites_to_process = [Site.objects.get(id=selected_site_id)]
                selected_site = Site.objects.get(id=selected_site_id)
                selected_site_name = selected_site.name
                selected_site_color = selected_site.color
            else:
                sites_to_process = Site.objects.filter(obsoleted=False)
                selected_site_name = '全拠点'
                selected_site_color = ''
        except Site.DoesNotExist:
            selected_site_name = '全拠点'
            selected_site_color = ''
            sites_to_process = Site.objects.filter(obsoleted=False)

        discrepancies = []

        for site in sites_to_process:
            transaction_aggregates = Transaction.objects.filter(
                Q(source_site_id=site.id) | Q(dest_site_id=site.id)
            ).annotate(
                adjusted_quantity=Case(
                    When(type='I', then=F('quantity')),
                    When(type='O', then=-F('quantity')),
                    When(type='T', source_site_id=site.id, then=-F('quantity')),
                    When(type='T', dest_site_id=site.id, then=F('quantity')),
                    default=Value(0),
                    output_field=IntegerField())
            ).values('name_id', 'lot', 'source_site_id', 'dest_site_id').annotate(
                total_quantity=Sum('adjusted_quantity'))

            inventory_items = InventoryItem.objects.filter(
                site_id=site.id) if site else InventoryItem.objects.all()

            for item in inventory_items:
                sum_quantities = 0
                for ta in transaction_aggregates:
                    if ta['lot'] == item.lot and ta['name_id'] == item.name_id and (
                            ta['source_site_id'] == item.site_id or ta['dest_site_id'] == item.site_id):
                        sum_quantities += ta['total_quantity']

                if sum_quantities != item.quantity:
                    discrepancies.append({
                        'id': item.id,
                        'name': item.name,
                        'site': item.site,
                        'unit': item.unit,
                        'lot': item.lot,
                        'inventory_quantity': item.quantity,
                        'transaction_quantity': sum_quantities,
                        'difference': item.quantity - sum_quantities
                    })

        context = {'discrepancies': discrepancies,
                   'sites': Site.objects.filter(obsoleted=False),
                   'selected_site_name': selected_site_name,
                   'selected_site_color': selected_site_color
                   }

        return render(request, 'inventory/validate_inventory.html', context)


class InventoryInspectionView(LoginRequiredMixin, View):
    def get(self, request):
        site_id = request.GET.get('site_id')
        context = self.get_context_data(site_id)
        return render(request, 'inventory/inventory_inspection.html', context)

    def post(self, request):
        action = request.POST.get('action')
        if action == 'save':
            return self.save_inspection(request)
        else:
            return self.handle_invalid_form(request)

    def save_inspection(self, request):
        site_id = request.COOKIES.get('selectedSite')

        if not site_id:
            messages.error(request,
                           "拠点が選択されていません。拠点を選択してから再度お試しください。")
            return redirect('inspection')

        try:
            site = Site.objects.get(id=site_id)
        except Site.DoesNotExist:
            messages.error(request, f"ID {site_id} の拠点が見つかりません。")
            return redirect('inspection')

        inspection_data = []
        current_time = timezone.now()

        for key, value in request.POST.items():
            if key.startswith('manual_input_'):
                drug_id = key.split('_')[-1]

                try:
                    drug = Drug.objects.get(id=drug_id)
                except Drug.DoesNotExist:
                    messages.error(request, f"ID {drug_id} の医薬品が見つかりません。")
                    return redirect('inspection')

                expected_quantity = request.POST.get(f'data_{drug_id}')
                actual_quantity = value

                if not expected_quantity or not actual_quantity:
                    messages.error(request, f"医薬品 {drug.name} の数量が入力されていません。")
                    return redirect('inspection')

                try:
                    discrepancy = int(expected_quantity) - int(actual_quantity)

                    inspection_data.append({
                        'drug_id': drug.id,
                        'drug_name': drug.name,
                        'expected_quantity': int(expected_quantity),
                        'actual_quantity': int(actual_quantity),
                        'discrepancy': discrepancy
                    })
                except ValueError:
                    messages.error(request, f"医薬品 {drug.name} の数量に無効な値が入力されました。")
                    return redirect('inspection')

        try:
            inspection = InventoryInspection(
                site=site,
                data=inspection_data,
                user=request.user,
                date_created=current_time
            )
            inspection.save()
            messages.success(request, "在庫確認が正常に保存されました。")
        except Exception as e:
            messages.error(request, f"在庫確認の保存中にエラーが発生しました: {str(e)}")

        return redirect('inspection')

    def handle_invalid_form(self, request):
        site_id = request.POST.get('site_id')
        context = self.get_context_data(site_id)
        context['error'] = "Invalid form submission. Please check your inputs."
        return render(request, 'inventory/inventory_inspection.html', context)

    def get_context_data(self, site_id):
        try:
            selected_site_id = site_id or self.request.COOKIES.get('selectedSite')
            selected_site = Site.objects.get(id=selected_site_id) if selected_site_id else None
            selected_site_name = selected_site.name if selected_site else '全拠点'
            selected_site_color = selected_site.color if selected_site else ''
        except Site.DoesNotExist:
            selected_site_name = '全拠点'
            selected_site_color = ''

        items_filter = {}
        if selected_site:
            items_filter['site_id'] = selected_site.id
            obsolete_drug_ids = DrugObsoleteBySite.objects.filter(
                site=selected_site,
                is_obsolete=True
            ).values_list('drug_id', flat=True)
        else:
            obsolete_drug_ids = []

        drug_quantities = InventoryItem.objects.filter(
            **items_filter,
            name__obsoleted=False  # Keep this existing condition
        ).exclude(
            name_id__in=obsolete_drug_ids
        ).values(
            'name', 'name__name', 'name__category__name', 'name__category__id', 'name__id'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by(
            'name__category__id', 'name__id'
        )

        inventory_by_category = {}
        for item in drug_quantities:
            category = item['name__category__name']
            if category not in inventory_by_category:
                inventory_by_category[category] = []
            inventory_by_category[category].append({
                'id': item['name'],
                'name': item['name__name'],
                'quantity': item['total_quantity']
            })

        # Fetch all inspections for the selected site
        all_inspections = InventoryInspection.objects.filter(site=selected_site).order_by('-date_created')

        # Dictionary to hold historical data per drug
        drug_histories = {}

        if all_inspections:
            for inspection in all_inspections:
                for record in inspection.data:
                    record['time'] = inspection.date_created.strftime('%Y-%m-%d %H:%M:%S')
                    record['username'] = inspection.user.username

                    # Group by drug name
                    if record['drug_name'] not in drug_histories:
                        drug_histories[record['drug_name']] = []
                    drug_histories[record['drug_name']].append(record)

        # Get the latest inspection time and username from the last inspection record
        latest_inspection = all_inspections.first()
        latest_inspection_time = latest_inspection.date_created.strftime(
            '%m-%d %H:%M') if latest_inspection else None
        latest_inspection_username = latest_inspection.user if latest_inspection else None
        latest_inventory_adjusted = latest_inspection.inventory_adjusted if latest_inspection else None

        categories = Drug.objects.filter(obsoleted=False).values('category__id', 'category__name').distinct()

        return {
            'inventory_by_category': inventory_by_category,
            'selected_site_name': selected_site_name,
            'selected_site_color': selected_site_color,
            'selected_site_id': selected_site_id,
            'inspector': latest_inspection_username,
            'latest_inspection_time': latest_inspection_time,
            'categories': categories,
            'drug_histories': drug_histories,
            'inventory_adjusted': latest_inventory_adjusted
        }


class InventoryInspectionHistoryView(LoginRequiredMixin, View):
    def get(self, request):
        site_id = request.GET.get('site_id')
        context = self.get_context_data(site_id)
        return render(request, 'inventory/inspection_history.html', context)

    def get_context_data(self, site_id):
        try:
            selected_site_id = site_id or self.request.COOKIES.get('selectedSite')
            selected_site = Site.objects.get(id=selected_site_id) if selected_site_id else None
            selected_site_name = selected_site.name if selected_site else '全拠点'
            selected_site_color = selected_site.color if selected_site else ''
        except Site.DoesNotExist:
            selected_site_name = '全拠点'
            selected_site_color = ''

        # Fetch all inspections for the selected site
        all_inspections = InventoryInspection.objects.filter(site=selected_site).order_by('-date_created')

        # Group inspection records by date
        historical_records = {}
        for inspection in all_inspections:
            inspection_time = inspection.date_created
            if inspection_time not in historical_records:
                historical_records[inspection_time] = []
            for record in inspection.data:
                if inspection.user.last_name:
                    record['username'] = f"{inspection.user.last_name} {inspection.user.first_name}"
                else:
                    record['username'] = inspection.user.username
                historical_records[inspection_time].append(record)

        return {
            'historical_records': historical_records,
            'selected_site_name': selected_site_name,
            'selected_site_color': selected_site_color,
            'selected_site_id': selected_site_id
        }


class ExportInventoryInspectionView(LoginRequiredMixin, View):
    def get(self, request):
        site_id = request.GET.get('site_id')

        try:
            selected_site_id = site_id or request.COOKIES.get('selectedSite')
            selected_site = Site.objects.get(id=selected_site_id) if selected_site_id else None
            selected_site_name = selected_site.name if selected_site else '全拠点'
        except Site.DoesNotExist:
            selected_site = None
            selected_site_name = '全拠点'

        # Fetch all inspections for the selected site
        all_inspections = InventoryInspection.objects.filter(site=selected_site).order_by('-date_created')

        if not all_inspections.exists():
            return redirect('inspection')

        # Create a workbook and select the active sheet
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        # Generate timestamp for filename
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')

        for inspection in all_inspections:
            # Create a new sheet for each inspection with date and time as the sheet name
            sheet_name = inspection.date_created.strftime('%Y-%m-%d %H-%M-%S')
            sheet = wb.create_sheet(title=sheet_name)

            # Write headers
            headers = ['医薬品名', '当時のシステム在庫数', '保存した実数', '在庫差異', '検査者']
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            # Write data
            inspector = f"{inspection.user.last_name} {inspection.user.first_name}" if inspection.user.last_name else inspection.user.username

            for row, record in enumerate(inspection.data, start=2):
                sheet.cell(row=row, column=1, value=record.get('drug_name', ''))
                sheet.cell(row=row, column=2, value=record.get('expected_quantity', ''))
                sheet.cell(row=row, column=3, value=record.get('actual_quantity', ''))
                sheet.cell(row=row, column=4, value=record.get('discrepancy', ''))
                sheet.cell(row=row, column=5, value=inspector)

        # Save the workbook to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create filename
        filename = f"実数カウント履歴_{selected_site_name}_{timestamp}.xlsx"
        encoded_filename = quote(filename)

        # Prepare the response
        response = HttpResponse(buffer.getvalue(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'

        return response


class InventoryBatchAdjustmentView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            site_id = request.GET.get('site_id') or request.COOKIES.get('selectedSite')

            if site_id:
                site = Site.objects.get(id=site_id)
                inspections = InventoryInspection.objects.filter(site=site)
            else:
                site = None
                inspections = InventoryInspection.objects.all()

            latest_inspection = inspections.order_by('-date_created').first()
            if not latest_inspection:
                context = {
                    'discrepancies': [],
                    'inspection_time': None,
                    'selected_site_name': site.name if site else '全拠点',
                    'selected_site_id': site.id if site else None,
                    'selected_site_color': site.color if site else None,
                    'no_data': True
                }
                return render(request, 'inventory/batch_adjustment.html', context)

            discrepancies = []
            zero_inventory_drugs = []

            for record in latest_inspection.data:
                if record['discrepancy'] != 0:
                    lot_numbers = InventoryItem.objects.filter(name_id=record['drug_id'], quantity__gt=0)
                    if site:
                        lot_numbers = lot_numbers.filter(site=site)

                    lot_info = []
                    has_positive_inventory = False

                    for lot in lot_numbers:
                        if lot.quantity > 0:
                            has_positive_inventory = True
                        lot_info.append({
                            'lot': lot.lot,
                            'quantity': lot.quantity,
                        })

                    if has_positive_inventory:
                        discrepancies.append({
                            'drug_id': record['drug_id'],
                            'drug_name': record['drug_name'],
                            'expected_quantity': record['expected_quantity'],
                            'actual_quantity': record['actual_quantity'],
                            'discrepancy': record['discrepancy'],
                            'lot_info': lot_info
                        })
                    else:
                        zero_inventory_drugs.append({
                            'drug_id': record['drug_id'],
                            'drug_name': record['drug_name']
                        })

            context = {
                'discrepancies': discrepancies,
                'zero_inventory_drugs': zero_inventory_drugs,
                'inventory_adjusted': latest_inspection.inventory_adjusted,
                'inspection_time': latest_inspection.date_created,
                'selected_site_name': site.name if site else '全拠点',
                'selected_site_id': site.id if site else None,
                'selected_site_color': site.color if site else None,
                'no_data': False
            }
            return render(request, 'inventory/batch_adjustment.html', context)

        except Exception as e:
            context = {
                'discrepancies': [],
                'inspection_time': None,
                'selected_site_name': site.name if site else '全拠点' if 'site' in locals() else '全拠点',
                'selected_site_id': site.id if site else None if 'site' in locals() else None,
                'selected_site_color': site.color if site else None,
                'no_data': True,
                'error_message': f"エラーが発生しました: {str(e)}"
            }
            return render(request, 'inventory/batch_adjustment.html', context)

    def post(self, request):
        try:
            site_id = request.COOKIES.get('selectedSite')
            site = Site.objects.get(id=site_id) if site_id else None

            with transaction.atomic():
                latest_inspection = InventoryInspection.objects.filter(site=site).order_by(
                    '-date_created').first()

                for key, value in request.POST.items():
                    if key.startswith('corrected_quantity_'):
                        parts = key.split('_')
                        drug_id = parts[2]
                        lot_index = parts[3]
                        corrected_quantity = int(value)
                        lot_number = request.POST.get(f'lot_number_{drug_id}_{lot_index}')

                        if not lot_number:
                            continue

                        drug = Drug.objects.get(id=drug_id)

                        inventory_item_query = InventoryItem.objects.filter(
                            name=drug,
                            lot=lot_number,
                            quantity__gt=0
                        )

                        if site:
                            inventory_item_query = inventory_item_query.filter(site=site)

                        inventory_item = inventory_item_query.first()

                        if not inventory_item:
                            messages.error(request,
                                           f"在庫項目が見つかりませんでした: {drug.name} (製造番号: {lot_number})")
                            continue

                        inventory_item.quantity = corrected_quantity
                        inventory_item.save()

                        Transaction.objects.create(
                            name=drug,
                            lot=lot_number,
                            source_site=site,
                            quantity=corrected_quantity,
                            expire_date=inventory_item.expire_date,
                            unit=drug.unit,
                            type=Transaction.Type.ADJUST,
                            user=request.user,
                            date_created=timezone.now()
                        )

                # Mark the latest inspection as adjusted
                if latest_inspection:
                    latest_inspection.inventory_adjusted = True
                    latest_inspection.save()

                messages.success(request, "在庫が正常に更新されました。")
                return redirect('inspection')

        except Exception as e:
            context = {
                'error_message': f"保存中にエラーが発生しました: {str(e)}",
                'discrepancies': [],
                'inspection_time': None,
                'selected_site_name': site.name if site else '全拠点',
                'selected_site_id': site.id if site else None,
                'no_data': False
            }
            return render(request, 'inventory/batch_adjustment.html', context)


def filtered_drugs(request):
    category_id = request.GET.get('categoryId')

    # Define the custom sort order annotation
    custom_sort_order = Case(
        When(kana__regex=r'^[ぁ-ん]', then=Value(1)),
        When(kana__regex=r'^[a-zA-Z]', then=Value(2)),
        default=Value(3),
        output_field=IntegerField(),
    )

    if category_id in ['', None]:
        drugs = Drug.objects.filter(obsoleted=False).annotate(
            custom_sort_order=custom_sort_order
        ).order_by(
            'custom_sort_order',
            'kana'
        ).values('id', 'name')
    else:
        drugs = Drug.objects.filter(category_id=category_id, obsoleted=False).annotate(
            custom_sort_order=custom_sort_order
        ).order_by(
            'custom_sort_order',
            'kana'
        ).values('id', 'name')

    return JsonResponse(list(drugs), safe=False)


def get_filtered_lots(request):
    drug_id = request.GET.get('drugId')
    site_id = request.GET.get('siteId')
    if drug_id in ['', None] or site_id in ['', None]:
        lots = []
    else:
        lots = InventoryItem.objects.filter(name_id=drug_id, site_id=site_id,
                                            quantity__gt=0).values('id', 'lot', 'expire_date')
    return JsonResponse(list(lots), safe=False)


def databar_parser(request):
    scanned_data = request.GET.get('scanned_data')
    return analyzer(scanned_data)


class ExportTransCSV(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        selected_site_id = ''
        try:
            selected_site_id = request.COOKIES.get('selectedSite')
            if selected_site_id is None:
                selected_site_name = '全拠点'
            else:
                selected_site_name = Site.objects.get(id=selected_site_id).name
        except Site.DoesNotExist:
            selected_site_name = '全拠点'

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"入出庫履歴_{selected_site_name}_{timestamp}.csv"
        encoded_filename = quote(filename)

        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}'},
        )

        writer = csv.writer(response)
        generate_transactions_csv(writer, selected_site_id)

        return response


@staff_member_required
def populate_safety_stock(request):
    if request.method == "POST":
        # Default value for min_stock
        default_min_stock = 1

        # Get all drugs and sites
        drugs = Drug.objects.all()
        sites = Site.objects.all()

        # Loop through drugs and sites for both operations
        for drug in drugs:
            for site in sites:
                # Update or create DrugObsoleteStatus
                DrugObsoleteBySite.objects.update_or_create(
                    drug=drug,
                    site=site,
                    defaults={'is_obsolete': False}
                )

                # Get or create StockSafety (renamed from SafetyStock)
                SafetyStock.objects.get_or_create(
                    drug=drug,
                    site=site,
                    defaults={'min_stock': default_min_stock}
                )

        return redirect('safety-stock-matrix')

    return render(request, 'inventory/populate_safety_stock.html')


@staff_member_required
def populate_kana(request):
    if request.method == "POST":
        kks = kakasi()
        kks.setMode("J", "H")  # Kanji to Hiragana
        kks.setMode('K', 'H')  # Katanaga to Hiragana
        conv = kks.getConverter()

        drugs = Drug.objects.all()
        updated_count = 0

        for drug in drugs:
            if drug.name:
                # Convert Kanji to Kana
                drug.kana = conv.do(drug.name)
                drug.save()
                updated_count += 1
        return redirect('drugs')
    return render(request, 'inventory/populate_kana.html')


class SafetyStockMatrixView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        safety_stock_form_set = modelformset_factory(SafetyStock, form=SafetyStockForm, extra=0)
        queryset = SafetyStock.objects.select_related('drug', 'site').filter(drug__obsoleted=False,
                                                                             site__obsoleted=False)
        formset = safety_stock_form_set(queryset=queryset)

        drugs = Drug.objects.filter(obsoleted=False).prefetch_related('safetystock_set')
        sites = Site.objects.filter(obsoleted=False)

        matrix = []
        form_dict = {(form.instance.drug_id, form.instance.site_id): form for form in formset}

        for drug in drugs:
            row = {'drug': drug, 'forms': []}
            for site in sites:
                form = form_dict.get((drug.id, site.id))
                if not form:
                    stock = SafetyStock(drug=drug, site=site)
                    form = SafetyStockForm(instance=stock)
                row['forms'].append(form)
            matrix.append(row)

        context = {
            'matrix': matrix,
            'sites': sites,
            'formset': formset
        }
        return render(request, 'inventory/safety_stock_matrix.html', context)

    def post(self, request, *args, **kwargs):
        safety_stock_form_set = modelformset_factory(SafetyStock, form=SafetyStockForm, extra=0)
        queryset = SafetyStock.objects.select_related('drug', 'site').filter(drug__obsoleted=False,
                                                                             site__obsoleted=False)
        formset = safety_stock_form_set(request.POST, queryset=queryset)

        if formset.is_valid():
            formset.save()
            return redirect('safety-stock-matrix')
        return self.get(request)


class ExportDrugMasterView(LoginRequiredMixin, View):
    def get(self, request):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"医薬品マスター_{timestamp}.csv"
        encoded_filename = quote(filename)

        dataset = DrugMasterResource().export()
        response = HttpResponse(
            dataset.csv,
            content_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}'}
        )
        return response


class ImportDrugMasterView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        if request.FILES:
            dataset = Dataset()
            new_file = request.FILES['myfile']

            try:
                imported_data = dataset.load(new_file.read().decode('utf-8'), format='csv')
            except UnicodeDecodeError as e:
                context = {'errors': [f"File encoding error: {str(e)}"]}
                return render(request, 'inventory/import_drug_master.html', context)
            except Exception as e:
                context = {'errors': [f"Unexpected error while loading the file: {str(e)}"]}
                return render(request, 'inventory/import_drug_master.html', context)
            resource = DrugMasterResource()
            # Perform dry run to test data
            result = resource.import_data(dataset, dry_run=True, collect_failed_rows=True)

            if result.has_errors():
                error_messages = []
                for error in result.row_errors():
                    # Collect row number and error
                    for e in error[1]:
                        error_message = f"Row {error[0]}: {str(e.error)}"
                        error_messages.append(error_message)

                context = {'errors': error_messages}
                return render(request, 'inventory/import_drug_master.html', context)
            else:
                # Perform the actual import
                resource.import_data(dataset, dry_run=False)
                return redirect('drugs')

        return render(request, 'inventory/import_drug_master.html')

    def get(self, request, *args, **kwargs):
        # Handles the GET requests; you might want to show an upload form here
        return render(request, 'inventory/import_drug_master.html')


@requires_csrf_token
def csrf_failure(request, reason=""):
    return render(request, 'inventory/errors/csrf_failure.html', {'reason': reason})


def handler400(request, exception):
    return render(request, 'inventory/errors/400.html', status=400)


def handler403(request, exception):
    return render(request, 'inventory/errors/403.html', status=403)


def handler404(request, exception):
    return render(request, 'inventory/errors/404.html', status=404)


def handler500(request):
    return render(request, 'inventory/errors/500.html', status=500)
